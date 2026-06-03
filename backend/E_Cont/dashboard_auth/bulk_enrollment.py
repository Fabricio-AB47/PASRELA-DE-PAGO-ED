from __future__ import annotations

import base64
import re
from dataclasses import dataclass
from io import BytesIO
from typing import Any

from django.db import connection
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .inscription_certificate import (
    build_certificate_payload,
    create_stored_certificate_record,
    send_certificate_email,
)
from .payments import PaymentGatewayError, create_mass_matriculation_and_credentials
from .student_registration import (
    RegisteredUserExistsError,
    USER_REGISTERED_MESSAGE,
    ensure_user_is_not_registered,
)


class BulkEnrollmentError(Exception):
    pass


@dataclass
class InMemoryExcelUpload:
    name: str
    content: bytes

    @property
    def size(self) -> int:
        return len(self.content)

    def seek(self, _position: int) -> None:
        return None

    def read(self) -> bytes:
        return self.content


TEMPLATE_FILE_NAME = 'plantilla_matricula_masiva.xlsx'
TEMPLATE_HEADERS = [
    'Nombres',
    'Apellidos',
    'Cédula',
    'Correo',
    'Número de celular',
    'Ciudad',
    'Dirección',
]
MAX_UPLOAD_BYTES = 5 * 1024 * 1024
MAX_BULK_ROWS = 500
MAX_STUDENT_SELECTION = 500
DEFAULT_EMPTY_LOCATION = 'No registrado'


HEADER_ALIASES = {
    'nombres': 'nombres',
    'nombre': 'nombres',
    'nombre completo': 'nombre_completo',
    'apellidos': 'apellidos',
    'apellido': 'apellidos',
    'cedula': 'cedula',
    'cédula': 'cedula',
    'identificacion': 'cedula',
    'identificación': 'cedula',
    'correo': 'email',
    'email': 'email',
    'correo electronico': 'email',
    'correo electrónico': 'email',
    'numero de celular': 'telefono',
    'número de celular': 'telefono',
    'celular': 'telefono',
    'telefono': 'telefono',
    'teléfono': 'telefono',
    'ciudad': 'localidad',
    'localidad': 'localidad',
    'direccion': 'direccion',
    'dirección': 'direccion',
}


def build_bulk_enrollment_template() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Carga'
    instructions = workbook.create_sheet('Instrucciones')
    example = workbook.create_sheet('Ejemplo')

    header_fill = PatternFill('solid', fgColor='9B0E0E')
    header_font = Font(color='FFFFFF', bold=True)
    border = Border(bottom=Side(style='thin', color='D9D9D9'))

    for column_index, header in enumerate(TEMPLATE_HEADERS, start=1):
        cell = sheet.cell(row=1, column=column_index, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
        sheet.column_dimensions[cell.column_letter].width = max(len(header) + 8, 20)

    for column in ('C', 'E'):
        sheet[f'{column}2'].number_format = '@'
        sheet.column_dimensions[column].width = 22

    instructions['A1'] = 'Plantilla válida para matrícula masiva'
    instructions['A1'].font = Font(bold=True, size=14, color='9B0E0E')
    instructions['A3'] = 'Columnas obligatorias: Nombres, Apellidos, Cédula, Correo, Número de celular.'
    instructions['A4'] = 'Columnas opcionales: Ciudad, Dirección.'
    instructions['A5'] = 'No cambies los nombres de las columnas.'
    instructions.column_dimensions['A'].width = 90

    for column_index, header in enumerate(TEMPLATE_HEADERS, start=1):
        cell = example.cell(row=1, column=column_index, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
        example.column_dimensions[cell.column_letter].width = max(len(header) + 8, 22)

    example.append(
        [
            'Juan Carlos',
            'Recalde Romo',
            '0012345678',
            'juan.recalde@example.com',
            '0999999999',
            'Quito',
            'Av. Principal 123',
        ]
    )
    for row in example.iter_rows(min_row=2, max_row=2, min_col=1, max_col=len(TEMPLATE_HEADERS)):
        for cell in row:
            cell.border = border
    example['C2'].number_format = '@'
    example['E2'].number_format = '@'

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def excel_upload_from_json(payload: dict[str, Any]) -> InMemoryExcelUpload:
    excel_payload = payload.get('excel') if isinstance(payload, dict) else None
    if not isinstance(excel_payload, dict):
        raise BulkEnrollmentError('Debes enviar el archivo Excel en el campo excel.')

    file_name = str(excel_payload.get('name') or '').strip()
    content_base64 = str(excel_payload.get('content_base64') or '').strip()
    if not file_name:
        raise BulkEnrollmentError('El archivo Excel debe incluir nombre.')
    if not file_name.lower().endswith('.xlsx'):
        raise BulkEnrollmentError('El archivo debe estar en formato .xlsx.')
    if not content_base64:
        raise BulkEnrollmentError('El archivo Excel esta vacio.')

    try:
        content = base64.b64decode(content_base64, validate=True)
    except (ValueError, TypeError) as exc:
        raise BulkEnrollmentError('El contenido del Excel no es base64 valido.') from exc

    upload = InMemoryExcelUpload(name=file_name, content=content)
    if upload.size > MAX_UPLOAD_BYTES:
        raise BulkEnrollmentError('El archivo excede el tamano maximo permitido para carga masiva.')
    return upload


def process_bulk_enrollment_excel(uploaded_file: Any, defaults: dict[str, Any]) -> dict[str, Any]:
    if not uploaded_file:
        raise BulkEnrollmentError('Debes seleccionar un archivo Excel para procesar.')

    file_name = str(getattr(uploaded_file, 'name', '') or '').strip()
    if not file_name.lower().endswith('.xlsx'):
        raise BulkEnrollmentError('El archivo debe estar en formato .xlsx.')

    file_size = int(getattr(uploaded_file, 'size', 0) or 0)
    if file_size > MAX_UPLOAD_BYTES:
        raise BulkEnrollmentError('El archivo excede el tamano maximo permitido para carga masiva.')

    clean_defaults = _clean_defaults(defaults)
    rows = _read_excel_rows(uploaded_file)
    if not rows:
        raise BulkEnrollmentError('El archivo no contiene filas para procesar.')
    if len(rows) > MAX_BULK_ROWS:
        raise BulkEnrollmentError(f'La carga masiva permite hasta {MAX_BULK_ROWS} filas por archivo.')

    results: list[dict[str, Any]] = []
    for row in rows:
        try:
            payload = _payload_from_row(row, clean_defaults)
            try:
                ensure_user_is_not_registered(
                    payload['cedula'],
                    cod_anio_basica=payload['cod_anio_basica'],
                    codigo_materia=payload['codigo_materia'],
                    codigo_periodo=payload['codigo_periodo'],
                )
            except RegisteredUserExistsError:
                results.append(
                    {
                        'ok': False,
                        'fila': row['fila'],
                        'nombre': payload['nombre'],
                        'cedula': payload['cedula'],
                        'email': payload['email'],
                        'registered': True,
                        'message': USER_REGISTERED_MESSAGE,
                    }
                )
                continue
            results.append(_process_matriculation_payload(payload, row['fila'], certificate_source='matricula_masiva'))
        except Exception as exc:
            results.append(
                {
                    'ok': False,
                    'fila': row.get('fila'),
                    'nombre': _row_full_name(row),
                    'cedula': row.get('cedula', ''),
                    'email': row.get('email', ''),
                    'message': str(exc),
                }
            )

    successful = sum(1 for item in results if item['ok'])
    failed = len(results) - successful
    return {
        'total': len(results),
        'exitosos': successful,
        'fallidos': failed,
        'results': results,
    }


def list_academic_enrollment_students(search: Any = '', limit: Any = 200) -> list[dict[str, str]]:
    clean_search = _clean_text(search)
    max_results = min(max(_safe_int(limit, default=200), 1), MAX_STUDENT_SELECTION)
    params: list[Any] = []
    filters = [
        "NULLIF(LTRIM(RTRIM(ISNULL(Apellidos_nombre, ''))), '') IS NOT NULL",
    ]

    if clean_search:
        like_value = f'%{clean_search}%'
        filters.append(
            """
            (
                LTRIM(RTRIM(ISNULL(Apellidos_nombre, ''))) LIKE %s
                OR REPLACE(REPLACE(LTRIM(RTRIM(ISNULL(Cedula_Est, ''))), '-', ''), ' ', '') LIKE %s
                OR LTRIM(RTRIM(ISNULL(CAST(Cedula AS varchar(20)), ''))) LIKE %s
                OR LTRIM(RTRIM(ISNULL(correo, ''))) LIKE %s
            )
            """
        )
        params.extend([like_value, like_value, like_value, like_value])

    query = f"""
        SELECT TOP ({max_results})
            CAST(codigo_estud AS varchar(50)) AS codigo_estud,
            LTRIM(RTRIM(ISNULL(Apellidos_nombre, ''))) AS nombre,
            LTRIM(RTRIM(COALESCE(NULLIF(Cedula_Est, ''), CAST(Cedula AS varchar(20)), ''))) AS cedula,
            LTRIM(RTRIM(ISNULL(correo, ''))) AS email,
            LTRIM(RTRIM(ISNULL(telefono, ''))) AS telefono,
            LTRIM(RTRIM(ISNULL(movil, ''))) AS movil,
            LTRIM(RTRIM(ISNULL(ciudad, ''))) AS localidad,
            LTRIM(RTRIM(ISNULL(calle_principal, ''))) AS direccion
        FROM dbo.DATOS_ESTUD
        WHERE {' AND '.join(filters)}
        ORDER BY Apellidos_nombre ASC, codigo_estud DESC
    """
    return [_serialize_student_candidate(row) for row in _fetch_all(query, params)]


def process_selected_student_enrollment(payload: dict[str, Any]) -> dict[str, Any]:
    clean_defaults = _clean_defaults(payload)
    selected_ids = _selected_student_ids(payload)
    if not selected_ids:
        raise BulkEnrollmentError('Selecciona al menos un estudiante para matricular.')
    if len(selected_ids) > MAX_STUDENT_SELECTION:
        raise BulkEnrollmentError(f'La matrícula por selección permite hasta {MAX_STUDENT_SELECTION} estudiantes.')

    students_by_id = _fetch_students_by_ids(selected_ids)
    results: list[dict[str, Any]] = []
    for index, codigo_estud in enumerate(selected_ids, start=1):
        student = students_by_id.get(codigo_estud)
        if not student:
            results.append(
                {
                    'ok': False,
                    'fila': index,
                    'codigo_estud': codigo_estud,
                    'nombre': '',
                    'cedula': '',
                    'email': '',
                    'message': 'El estudiante seleccionado no existe en DATOS_ESTUD.',
                }
            )
            continue

        try:
            student_payload = _payload_from_student(student, clean_defaults)
            processed = _process_matriculation_payload(
                student_payload,
                index,
                certificate_source='matricula_academica',
            )
            processed['codigo_estud'] = codigo_estud
            results.append(processed)
        except Exception as exc:
            results.append(
                {
                    'ok': False,
                    'fila': index,
                    'codigo_estud': codigo_estud,
                    'nombre': student.get('nombre', ''),
                    'cedula': student.get('cedula', ''),
                    'email': student.get('email', ''),
                    'message': str(exc),
                }
            )

    successful = sum(1 for item in results if item['ok'])
    failed = len(results) - successful
    return {
        'total': len(results),
        'exitosos': successful,
        'fallidos': failed,
        'results': results,
    }


def _process_matriculation_payload(
    payload: dict[str, Any],
    row_label: Any,
    *,
    certificate_source: str,
) -> dict[str, Any]:
    result = create_mass_matriculation_and_credentials(payload)
    welcome_email_result = result.get('welcome_email_result', {})
    welcome_email_sent = bool(welcome_email_result.get('sent'))
    official_record = result.get('official_sync', {}).get('record') or {}
    certificate_payload = build_certificate_payload(payload, result, source=certificate_source)
    certificate_record = create_stored_certificate_record(certificate_payload)
    try:
        certificate_email_result = send_certificate_email(
            recipient_email=payload['email'],
            recipient_name=payload['nombre'],
            certificate_record=certificate_record,
        )
    except Exception as exc:
        certificate_email_result = {
            'sent': False,
            'message': f'Certificado generado, pero no fue posible enviarlo por correo: {str(exc)}',
            'filename': certificate_record.get('filename'),
        }

    return {
        'ok': True,
        'fila': row_label,
        'nombre': payload['nombre'],
        'cedula': payload['cedula'],
        'email': payload['email'],
        'matricula': result.get('matricula'),
        'codigo_materia': official_record.get('codigo_materia') or payload['codigo_materia'],
        'materia': official_record.get('materia') or payload.get('nombre_materia'),
        'welcome_email_sent': welcome_email_sent,
        'welcome_email_message': str(welcome_email_result.get('message') or ''),
        'microsoft365_ok': bool(result.get('microsoft365', {}).get('ok')),
        'certificate': certificate_record,
        'certificate_email_sent': bool(certificate_email_result.get('sent')),
        'certificate_email_message': str(certificate_email_result.get('message') or ''),
        'message': (
            'Procesado correctamente.'
            if welcome_email_sent
            else str(welcome_email_result.get('message') or 'Procesado, pero la bienvenida quedó pendiente.')
        ),
    }


def _selected_student_ids(payload: dict[str, Any]) -> list[str]:
    raw_values = payload.get('student_ids')
    if raw_values is None:
        raw_values = payload.get('selected_student_ids')
    if not isinstance(raw_values, list):
        raise BulkEnrollmentError('Debes enviar la lista de estudiantes seleccionados.')

    selected_ids: list[str] = []
    seen: set[str] = set()
    for value in raw_values:
        clean_value = _clean_text(value)
        if clean_value and clean_value not in seen:
            selected_ids.append(clean_value)
            seen.add(clean_value)
    return selected_ids


def _fetch_students_by_ids(student_ids: list[str]) -> dict[str, dict[str, str]]:
    if not student_ids:
        return {}

    placeholders = ', '.join(['%s'] * len(student_ids))
    query = f"""
        SELECT
            CAST(codigo_estud AS varchar(50)) AS codigo_estud,
            LTRIM(RTRIM(ISNULL(Apellidos_nombre, ''))) AS nombre,
            LTRIM(RTRIM(COALESCE(NULLIF(Cedula_Est, ''), CAST(Cedula AS varchar(20)), ''))) AS cedula,
            LTRIM(RTRIM(ISNULL(correo, ''))) AS email,
            LTRIM(RTRIM(ISNULL(telefono, ''))) AS telefono,
            LTRIM(RTRIM(ISNULL(movil, ''))) AS movil,
            LTRIM(RTRIM(ISNULL(ciudad, ''))) AS localidad,
            LTRIM(RTRIM(ISNULL(calle_principal, ''))) AS direccion
        FROM dbo.DATOS_ESTUD
        WHERE CAST(codigo_estud AS varchar(50)) IN ({placeholders})
    """
    return {
        str(row.get('codigo_estud') or '').strip(): _serialize_student_candidate(row)
        for row in _fetch_all(query, student_ids)
        if str(row.get('codigo_estud') or '').strip()
    }


def _payload_from_student(student: dict[str, str], defaults: dict[str, str]) -> dict[str, Any]:
    nombre = _clean_text(student.get('nombre'))
    cedula = re.sub(r'\D+', '', _clean_text(student.get('cedula')))
    email = _clean_text(student.get('email')).lower()
    telefono = _clean_text(student.get('telefono')) or _clean_text(student.get('movil'))
    if not nombre:
        raise PaymentGatewayError('Falta nombre del estudiante.')
    if not cedula:
        raise PaymentGatewayError('Falta cédula del estudiante.')
    if not email:
        raise PaymentGatewayError('Falta correo del estudiante.')
    if not telefono:
        raise PaymentGatewayError('Falta número de celular del estudiante.')

    direccion = _clean_text(student.get('direccion')) or DEFAULT_EMPTY_LOCATION
    localidad = _clean_text(student.get('localidad')) or DEFAULT_EMPTY_LOCATION
    course_name = _clean_text(defaults.get('nombre_materia'))
    descripcion = f'Matrícula académica del curso {course_name}' if course_name else 'Matrícula académica'

    return {
        'nombre': nombre,
        'cedula': cedula,
        'email': email,
        'telefono': telefono,
        'localidad': localidad,
        'direccion': direccion,
        'descripcion': descripcion,
        'nombre_materia': course_name,
        'carrera_num': defaults.get('carrera_num', ''),
        'cod_anio_basica': defaults['cod_anio_basica'],
        'codigo_materia': defaults['codigo_materia'],
        'codigo_periodo': defaults['codigo_periodo'],
        'estado_periodo': defaults.get('estado_periodo', ''),
        'data_treatment_accepted': True,
        'provider_payload': {
            'tipo': 'matricula_academica_sin_cargo',
            'nombre': nombre,
            'cedula': cedula,
            'email': email,
            'telefono': telefono,
            'localidad': localidad,
            'direccion': direccion,
            'descripcion': descripcion,
            'nombre_materia': course_name,
            'carrera_num': defaults.get('carrera_num', ''),
            'cod_anio_basica': defaults['cod_anio_basica'],
            'codigo_materia': defaults['codigo_materia'],
            'codigo_periodo': defaults['codigo_periodo'],
            'estado_periodo': defaults.get('estado_periodo', ''),
        },
    }


def _serialize_student_candidate(row: dict[str, Any]) -> dict[str, str]:
    telefono = _clean_text(row.get('telefono')) or _clean_text(row.get('movil'))
    return {
        'codigo_estud': _clean_text(row.get('codigo_estud')),
        'nombre': _clean_text(row.get('nombre')),
        'cedula': re.sub(r'\D+', '', _clean_text(row.get('cedula'))),
        'email': _clean_text(row.get('email')).lower(),
        'telefono': telefono,
        'localidad': _clean_text(row.get('localidad')),
        'direccion': _clean_text(row.get('direccion')),
    }


def _clean_defaults(defaults: dict[str, Any]) -> dict[str, str]:
    required = {
        'cod_anio_basica': 'Debes seleccionar la carrera para la matrícula.',
        'codigo_materia': 'Debes seleccionar el curso para la matrícula.',
        'codigo_periodo': 'Debes seleccionar el período para la matrícula.',
    }
    cleaned = {key: str(value or '').strip() for key, value in defaults.items()}
    for field, message in required.items():
        if not cleaned.get(field):
            raise BulkEnrollmentError(message)

    estado_periodo = cleaned.get('estado_periodo', '').lower()
    if estado_periodo and estado_periodo != 'activo':
        raise BulkEnrollmentError('El período seleccionado está inactivo.')
    return cleaned


def _read_excel_rows(uploaded_file: Any) -> list[dict[str, str]]:
    try:
        raw_content = uploaded_file.read()
        workbook = load_workbook(filename=BytesIO(raw_content), data_only=True)
    except Exception as exc:
        raise BulkEnrollmentError('No fue posible leer el archivo Excel. Verifica el formato.') from exc

    try:
        worksheet = workbook['Carga'] if 'Carga' in workbook.sheetnames else workbook.active
        header_map = _header_map(worksheet)
        rows: list[dict[str, str]] = []

        for row_index in range(2, worksheet.max_row + 1):
            row_data: dict[str, str] = {'fila': row_index}
            has_value = False
            for column_index, field_name in header_map.items():
                value = worksheet.cell(row=row_index, column=column_index).value
                text = _clean_text(value)
                if text:
                    has_value = True
                row_data[field_name] = text
            if has_value:
                rows.append(row_data)
        return rows
    finally:
        workbook.close()


def _header_map(worksheet: Any) -> dict[int, str]:
    mapping: dict[int, str] = {}
    for column_index in range(1, worksheet.max_column + 1):
        header = _normalize_header(worksheet.cell(row=1, column=column_index).value)
        field_name = HEADER_ALIASES.get(header)
        if field_name:
            mapping[column_index] = field_name

    required = {'cedula', 'email', 'telefono'}
    if not ({'nombre_completo'} <= set(mapping.values()) or {'nombres', 'apellidos'} <= set(mapping.values())):
        raise BulkEnrollmentError('El Excel debe incluir Nombre completo o Nombres y Apellidos.')
    missing = required - set(mapping.values())
    if missing:
        raise BulkEnrollmentError('Faltan columnas obligatorias: ' + ', '.join(sorted(missing)))
    return mapping


def _payload_from_row(row: dict[str, str], defaults: dict[str, str]) -> dict[str, Any]:
    nombre = _row_full_name(row)
    cedula = re.sub(r'\D+', '', _clean_text(row.get('cedula')))
    email = _clean_text(row.get('email')).lower()
    telefono = _clean_text(row.get('telefono'))
    if not nombre:
        raise PaymentGatewayError('Falta nombre del estudiante.')
    if not cedula:
        raise PaymentGatewayError('Falta cédula del estudiante.')
    if not email:
        raise PaymentGatewayError('Falta correo del estudiante.')
    if not telefono:
        raise PaymentGatewayError('Falta numero de celular del estudiante.')

    direccion = _clean_text(row.get('direccion')) or DEFAULT_EMPTY_LOCATION
    localidad = _clean_text(row.get('localidad')) or DEFAULT_EMPTY_LOCATION
    course_name = _clean_text(defaults.get('nombre_materia'))
    descripcion = f'Matrícula masiva del curso {course_name}' if course_name else 'Matrícula masiva'

    return {
        'nombre': nombre,
        'cedula': cedula,
        'email': email,
        'telefono': telefono,
        'localidad': localidad,
        'direccion': direccion,
        'descripcion': descripcion,
        'nombre_materia': course_name,
        'carrera_num': defaults.get('carrera_num', ''),
        'cod_anio_basica': defaults['cod_anio_basica'],
        'codigo_materia': defaults['codigo_materia'],
        'codigo_periodo': defaults['codigo_periodo'],
        'estado_periodo': defaults.get('estado_periodo', ''),
        'data_treatment_accepted': True,
        'provider_payload': {
            'tipo': 'matricula_masiva_sin_cargo',
            'nombre': nombre,
            'cedula': cedula,
            'email': email,
            'telefono': telefono,
            'localidad': localidad,
            'direccion': direccion,
            'descripcion': descripcion,
            'nombre_materia': course_name,
            'carrera_num': defaults.get('carrera_num', ''),
            'cod_anio_basica': defaults['cod_anio_basica'],
            'codigo_materia': defaults['codigo_materia'],
            'codigo_periodo': defaults['codigo_periodo'],
            'estado_periodo': defaults.get('estado_periodo', ''),
        },
    }


def _row_full_name(row: dict[str, str]) -> str:
    nombre_completo = _clean_text(row.get('nombre_completo'))
    if nombre_completo:
        return nombre_completo
    nombres = _clean_text(row.get('nombres'))
    apellidos = _clean_text(row.get('apellidos'))
    return _clean_text(f'{nombres} {apellidos}')


def _normalize_header(value: Any) -> str:
    return re.sub(r'\s+', ' ', str(value or '').strip().lower())


def _safe_int(value: Any, default: int = 0) -> int:
    try:
        return int(float(str(value or '').strip()))
    except (TypeError, ValueError):
        return default


def _fetch_all(query: str, params: list[Any]) -> list[dict[str, Any]]:
    with connection.cursor() as cursor:
        cursor.execute(query, params)
        columns = [column[0] for column in cursor.description]
        return [dict(zip(columns, row)) for row in cursor.fetchall()]


def _clean_text(value: Any) -> str:
    return re.sub(r'\s+', ' ', str(value or '').strip())
